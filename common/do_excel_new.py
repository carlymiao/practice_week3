# -*- coding: utf-8 -*-
# @Time    : 2019/3/17 12:27
# @Author  : Miao
# @Email   : 3266565259@qq.com
# @File    : do_excel.py

from openpyxl import load_workbook
from common.log import Log
from common.readconfig import ReadConfig
import traceback
import pandas as pd

from common import project_path


class DoExcel:
    """该类完成测试数据的读取和测试结果写回"""

    def __init__(self, file_name):
        self.file_name = file_name

    def read_tel(self):

        try:
            wb = load_workbook(self.file_name)
            sheet = wb['tel']
            tel = sheet.cell(1, 2).value
            Log().info('---读取电话号码{}'.format(tel))
        except Exception as e:
            Log().error('---读取电话号码失败{}'.format(e))
        wb.close()
        return tel

    def write_tel(self, old_tel):
        try:
            wb = load_workbook(self.file_name)
            sheet = wb['tel']

        except Exception as e:
            Log().error('---更新电话号码失败{}'.format(e))
        new_tel = old_tel + 1
        sheet.cell(1, 2).value = new_tel
        Log().info('---更新电话号码{}'.format(new_tel))
        wb.save(self.file_name)
        wb.close()

    def read_excel(self):
        """read模块把不同sheet页的用例根据配置文件设置全部取出，"""
        try:
            wb = load_workbook(self.file_name)
            Log().info('-----------------------------------------开始读取测试数据-----------------------------------------')
        except Exception as e:
            Log().error('---打开用例失败{}'.format(traceback.format_exc()))
            print(e)

        final_data = []
        case_amount = ReadConfig().case_amount()
        # column_name = ['CaseId', 'Module', 'Url', 'Method', 'Description', 'Param', 'ExpectedResult']
        # for row in range(2, sheet.max_row+1):
        #     row_data = {}
        #     for col, item in enumerate(column_name):
        #         row_data[item] =sheet.cell(row, col + 1).value
        #
        #     data_case.append(row_data)
        # wb.close()
        # print(data_case)
        # return data_case
        # 方法二
        for key in case_amount:  # 取到配置文件{'recharge':0,'Sheet1':[1,3,5]}，
            data_case = []
            sheet = wb[key]  # key代表表单名
            for row in range(2, sheet.max_row + 1):
                # print(row)
                row_data = {}
                row_data['CaseId'] = sheet.cell(row, 1).value
                row_data['Module'] = sheet.cell(row, 2).value
                row_data['Url'] = sheet.cell(row, 3).value
                row_data['Method'] = sheet.cell(row, 4).value
                row_data['Description'] = sheet.cell(row, 5).value
                # 方法二
                row_data['Param'] = eval(sheet.cell(row, 6).value)
                if row_data['Param'].get('mobilephone') == 'tel':
                    row_data['Param']['mobilephone'] = DoExcel(project_path.test_cases_path).read_tel()
                    DoExcel(project_path.test_cases_path).write_tel(DoExcel(project_path.test_cases_path).read_tel())
                row_data['Sql'] = eval(sheet.cell(row, 7).value)
                row_data['ExpectedResult'] = eval(sheet.cell(row, 8).value)
                data_case.append(row_data)


            # 此处获取配置文件配置用例,0代表执行全部用例
            if case_amount[key] == 0:
                final_data += data_case
            else:
                for i in case_amount[key]:
                    final_data.append(data_case[int(i) - 1])
        wb.close()
        Log.info('测试用例共{}条，分别为：{}'.format(len(final_data), final_data))
        Log.info('--------------------------------------------读取测试数据完毕-----------------------------------------')
        print(len(final_data), final_data)
        return final_data

    def write_excel(self, sheet_name, row, col, value):
        try:
            wb = load_workbook(self.file_name)
            sheet = wb[sheet_name]
            sheet.cell(row, col).value = '{}'.format(value)
            wb.save(self.file_name)
            wb.close()
        except Exception as e:
            Log().error('---写回测试结果失败{}'.format(e))


if __name__ == '__main__':
    t = DoExcel(project_path.test_cases_path).read_excel()
    # t = DoExcel(project_path.test_cases_path).read_excel('tel')
    # t = DoExcel(project_path.test_cases_path).read_excel('recharge', 'RECHARGE')
    # t.read_excel()
